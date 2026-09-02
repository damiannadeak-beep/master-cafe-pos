import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_interview_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instrumen Wawancara TA"

    # Gridlines view
    ws.views.sheetView[0].showGridLines = True

    # Color Palette (Modern Navy / Slate Teal theme)
    PRIMARY_COLOR = "1F4E78"    # Navy Dark Header
    HEADER_FILL = "2F5597"      # Table Header Fill
    SECTION_FILL = "D9E1F2"     # Section Header Fill
    ALT_ROW_FILL = "F2F5F9"     # Alternating Row Fill
    BORDER_COLOR = "D9D9D9"     # Light Gray Border

    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
    font_meta_label = Font(name="Calibri", size=10, bold=True)
    font_meta_val = Font(name="Calibri", size=10)
    font_section = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_code = Font(name="Calibri", size=10, bold=True, color="2F5597")

    # Fills
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_section = PatternFill(start_color=SECTION_FILL, end_color=SECTION_FILL, fill_type="solid")
    fill_alt = PatternFill(start_color=ALT_ROW_FILL, end_color=ALT_ROW_FILL, fill_type="solid")

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Borders
    thin_border_side = Side(border_style="thin", color=BORDER_COLOR)
    thick_bottom_side = Side(border_style="medium", color=PRIMARY_COLOR)
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_section = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)

    # 1. Title Banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "INSTRUMEN WAWANCARA ANALISIS KEBUTUHAN SISTEM INFORMASI POS"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Studi Kasus: Master Cafe (Bengkalis, Riau) - Pengumpulan Data Tugas Akhir (TA)"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # 2. Metadata Section
    metadata = [
        ("Nama Usaha / Objek Study", "Master Cafe", "Tanggal Wawancara", "____ / ____ / 2026"),
        ("Lokasi", "Bengkalis, Riau", "Narasumber / Responden", "____________________"),
        ("Pewawancara", "Damian (Mahasiswa Polbeng)", "Jabatan Responden", "Pemilik / Manajer Cafe")
    ]

    row_idx = 4
    for item in metadata:
        ws.cell(row=row_idx, column=1, value=item[0]).font = font_meta_label
        ws.cell(row=row_idx, column=2, value=item[1]).font = font_meta_val
        ws.cell(row=row_idx, column=4, value=item[2]).font = font_meta_label
        ws.cell(row=row_idx, column=5, value=item[3]).font = font_meta_val
        row_idx += 1

    row_idx += 1 # Empty row spacing

    # 3. Table Headers
    headers = [
        "No",
        "Kategori / Aspek Analisis",
        "Kode",
        "Pertanyaan Wawancara",
        "Relevansi & Fitur System (Analisis TA)",
        "Hasil Wawancara / Jawaban Narasumber",
        "Catatan & Tindak Lanjut"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    
    ws.row_dimensions[row_idx].height = 28

    # 4. Questions Data grouped by Sections
    sections_data = [
        {
            "category": "A. Profil & Operasional Usaha",
            "questions": [
                ("A1", "Jam berapa operasional cafe buka dan tutup setiap harinya?", "Konfigurasi Jam Operasional & Master Data Store"),
                ("A2", "Berapa jumlah meja/kapasitas tempat duduk yang tersedia saat ini?", "Kapasitas Data Master Meja (Table Layout Management)"),
                ("A3", "Berapa rata-rata jumlah transaksi atau pengunjung per hari (Weekday vs Weekend)?", "Proyeksi Beban Server & Transaksi Harian Sistem"),
                ("A4", "Berapa jumlah staf yang bekerja (Kasir, Pelayan, Barista, Dapur, Owner) dan bagaimana pembagian tugasnya?", "Perancangan Role Access (RBAC) & Multi-User Privileges")
            ]
        },
        {
            "category": "B. Alur Pemesanan (Ordering Process) & Manajemen Meja",
            "questions": [
                ("B1", "Bagaimana alur pemesanan saat ini? (Apakah pesan & bayar di kasir dulu, atau pesan di meja baru bayar di akhir?)", "Penentuan Alur Workflow POS (Pay-First vs Pay-Later)"),
                ("B2", "Apakah pelanggan sering berpindah meja atau menambah pesanan (add-on) saat berada di lokasi?", "Fitur Pindah Meja (Move Table) & Gabung/Tambah Pesanan"),
                ("B3", "Apakah cafe membutuhkan denah meja digital untuk memantau status meja (Terisi, Kosong, Reserved)?", "Fitur Live Table Management & Status Visual Meja"),
                ("B4", "Apakah tertarik menerapkan Pemesanan Mandiri via QR Code di Meja oleh pelanggan?", "Fitur Self-Service QR Ordering System")
            ]
        },
        {
            "category": "C. Alur Pembayaran & Kasir (Point of Sale)",
            "questions": [
                ("C1", "Apa saja metode pembayaran yang diterima saat ini? (Cash, QRIS, Transfer Bank, Debit, E-wallet)", "Modul Multi-Payment Method & Dynamic QRIS Integration"),
                ("C2", "Apakah ada kebutuhan Split Bill (satu meja/nota dibayar terpisah oleh beberapa orang)?", "Fitur Split Bill & Partial Payment Handler"),
                ("C3", "Apakah cafe mencetak struk fisik thermal, atau juga membutuhkan e-receipt (WhatsApp/Email)?", "Integrasi Thermal Printer & Digital Invoice Generator")
            ]
        },
        {
            "category": "D. Alur Dapur & Bar (Kitchen & Bar Workflow)",
            "questions": [
                ("D1", "Bagaimana pesanan dari kasir/pelayan diteruskan ke Barista (minuman) dan Dapur (makanan)?", "Workflow Order Routing & Ticket Dispatcher"),
                ("D2", "Apakah perlu pencetakan printer terpisah (Printer Bar khusus minuman & Printer Dapur khusus makanan)?", "Fitur Multi-Printer Router (Split Ticket by Category)"),
                ("D3", "Apakah sering terjadi kendala pesanan terlambat, salah buat, atau terlewat saat cafe sedang ramai?", "Fitur Kitchen Display System (KDS) & Timer Tracking Pesanan")
            ]
        },
        {
            "category": "E. Manajemen Stok & Bahan Baku (Inventory Management)",
            "questions": [
                ("E1", "Bagaimana pencatatan stok saat ini? (Hanya barang jadi atau sampai ke bahan baku mentah seperti kopi/susu/sirup)?", "Arsitektur Database Inventory & Stock Management"),
                ("E2", "Apakah cafe membutuhkan fitur Resep / Bill of Materials (BOM)? (Contoh: 1 Cappuccino otomatis potong stok 18g Kopi & 150ml Susu)", "Modul Recipe / BOM Auto-Deduction Inventory System"),
                ("E3", "Bagaimana proses pencatatan jika ada bahan baku yang kadaluarsa, rusak, atau terbuang (waste)?", "Fitur Stock Wastage & Adjustment Journal"),
                ("E4", "Siapa yang bertanggung jawab melakukan input restok / pengadaan bahan baku?", "Hak Akses Inventory Manager / Stock Opname Role")
            ]
        },
        {
            "category": "F. Promo, Diskon, & Pelanggan (Marketing & Loyalty)",
            "questions": [
                ("F1", "Apakah Master Cafe sering mengadakan promo diskon, paket bundling, atau promo hari tertentu?", "Engine Calculation Promo (Persentase, Nominal, Package Bundling)"),
                ("F2", "Apakah cafe mengumpulkan data pelanggan (Nama/No. WA) untuk sistem member/loyalty point?", "Modul Customer Relationship Management (CRM) & Loyalty Point"),
                ("F3", "Apakah perlu ada fitur Ulasan / Rating dari pelanggan terhadap makanan/pelayanan?", "Fitur Customer Feedback & Rating Analytics Module")
            ]
        },
        {
            "category": "G. Manajemen Shift Kasir & Rekonsiliasi Kas (Cash Control)",
            "questions": [
                ("G1", "Apakah menerapkan sistem Shift Kerja Kasir (Pagi & Malam)?", "Modul Shift Management & Cashier Session Tracking"),
                ("G2", "Bagaimana prosedur serah terima shift? (Apakah ada penginputan Modal Kas Awal & Hitung Uang Fisik Kas Akhir)?", "Fitur Cash Drawer Reconcile & Closing Shift Report"),
                ("G3", "Bagaimana kecocokan setoran uang tunai kasir dengan laporan penjualan harian?", "Laporan Audit Selisih Kas (Cash Variance Report)")
            ]
        },
        {
            "category": "H. Otorisasi Keamanan & Pembatalan Transaksi (Security & RBAC)",
            "questions": [
                ("H1", "Jika ada batal pesanan/salah input setelah nota dicetak, siapa yang berhak membatalkan (Void/Refund)?", "Fitur Supervisor PIN Authorization for Void/Refund"),
                ("H2", "Apakah kasir biasa boleh memberi diskon manual atau merubah harga item?", "Permission Control Diskon Manual & Edit Price"),
                ("H3", "Masalah kecurangan (fraud) atau selisih nota apa yang pernah diantisipasi manajemen?", "Audit Log Trail Transaksi & Anti-Fraud Security Logging")
            ]
        },
        {
            "category": "I. Pajak, Service Charge & Tipe Layanan",
            "questions": [
                ("I1", "Apakah cafe membebankan Pajak Resto (PB1 10%) atau Service Charge kepada pelanggan?", "Kalkulasi Subtotal, Tax (PB1), Service Charge & Grand Total"),
                ("I2", "Apakah harga menu sudah termasuk pajak (Inclusive) atau ditambah di akhir (Exclusive)?", "Setting Flag Tax Inclusive / Exclusive System"),
                ("I3", "Apakah ada perbedaan harga atau biaya packaging untuk Takeaway vs Dine-In?", "Dynamic Price Surcharge by Service Type (Dine-in/Takeaway)")
            ]
        },
        {
            "category": "J. Infrastruktur Perangkat & Kebutuhan Teknis (Hardware & System)",
            "questions": [
                ("J1", "Perangkat apa yang digunakan di kasir/dapur (PC Windows, Laptop, Tablet Android, HP)?", "Spesifikasi Minimal Compatibility Runtime Application"),
                ("J2", "Jenis koneksi printer thermal yang ada di lokasi (USB, Bluetooth, atau LAN)?", "Driver Support & Printing Service Module (ESC/POS Printer)"),
                ("J3", "Bagaimana kualitas jaringan internet di lokasi cafe? Apakah butuh sistem yang tetap bisa jalan saat internet mati?", "Desain Arsitektur Sync (Cloud vs Local First Database)")
            ]
        },
        {
            "category": "K. Laporan & Kebutuhan Manajemen (Reporting & Pain Points)",
            "questions": [
                ("K1", "Laporan apa saja yang paling krusial untuk Pemilik? (Penjualan Harian, Profit Bersih, Menu Terlaris, Stok Menipis)", "Desain Dashboard Analytics & Export PDF/Excel Reports"),
                ("K2", "Seberapa sering pemilik memantau operasional dan apakah butuh pemantauan jarak jauh via HP/Laptop?", "Fitur Owner Remote Dashboard & Web Monitoring Portal"),
                ("K3", "Apa kendala terbesar saat ini dan apa ekspektasi utama pemilik terhadap aplikasi baru ini?", "Formulasi Problem Statement & Key Performance Metric TA")
            ]
        }
    ]

    # Fill Table Content
    global_no = 1
    current_row = row_idx + 1

    for section in sections_data:
        # Write Section Header Row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        sec_cell = ws.cell(row=current_row, column=1, value=section["category"])
        sec_cell.font = font_section
        sec_cell.fill = fill_section
        sec_cell.alignment = align_left
        
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_section

        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Write Questions in Section
        for q_code, q_text, q_relevance in section["questions"]:
            is_alt = (global_no % 2 == 0)
            row_fill = fill_alt if is_alt else None

            c1 = ws.cell(row=current_row, column=1, value=global_no)
            c2 = ws.cell(row=current_row, column=2, value=section["category"].split(". ")[1])
            c3 = ws.cell(row=current_row, column=3, value=q_code)
            c4 = ws.cell(row=current_row, column=4, value=q_text)
            c5 = ws.cell(row=current_row, column=5, value=q_relevance)
            c6 = ws.cell(row=current_row, column=6, value="") # Answer placeholder
            c7 = ws.cell(row=current_row, column=7, value="") # Notes placeholder

            # Styling row cells
            c1.alignment = align_center
            c2.alignment = align_left
            c3.alignment = align_center
            c4.alignment = align_left
            c5.alignment = align_left
            c6.alignment = align_top_left
            c7.alignment = align_top_left

            c1.font = font_data
            c2.font = font_data
            c3.font = font_code
            c4.font = font_data
            c5.font = font_data

            for col_cell in (c1, c2, c3, c4, c5, c6, c7):
                col_cell.border = border_cell
                if row_fill:
                    col_cell.fill = row_fill

            ws.row_dimensions[current_row].height = 40
            global_no += 1
            current_row += 1

    # 5. Set Column Widths
    col_widths = {
        1: 6,    # No
        2: 24,   # Kategori
        3: 8,    # Kode
        4: 45,   # Pertanyaan Wawancara
        5: 35,   # Relevansi/Fitur System
        6: 45,   # Hasil Wawancara
        7: 25    # Catatan
    }

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Save output file
    output_path = r"c:\xampp\htdocs\angkringan-pos\Daftar_Pertanyaan_Wawancara_Master_Cafe_Bengkalis.xlsx"
    wb.save(output_path)
    print(f"Excel successfully created at: {output_path}")

if __name__ == "__main__":
    create_interview_excel()
